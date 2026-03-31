import openpyxl
import time
import sys
import re
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options as ChromeOptions
from webdriver_manager.chrome import ChromeDriverManager


def clean_number(raw):
    """Normalize a contact number: strip formatting, keep digits + optional leading '+'."""
    if raw is None:
        return None
        
    # Openpyxl often reads large numbers as floats (e.g. 919512900769.0)
    if isinstance(raw, float) and raw.is_integer():
        raw = int(raw)
        
    text = str(raw).strip()
    if not text:
        return None
        
    # Remove trailing '.0' just in case it arrived as a string
    if text.endswith('.0'):
        text = text[:-2]
        
    cleaned = re.sub(r'[\s\-\(\)\.]', '', text)
    if cleaned.startswith('+'):
        digits = '+' + re.sub(r'[^0-9]', '', cleaned[1:])
    else:
        digits = re.sub(r'[^0-9]', '', cleaned)
    return digits if len(digits) >= 7 else None


def find_contact_column(sheet):
    """Auto-detect the column with contact numbers from the header row."""
    keywords = ['contact', 'phone', 'mobile', 'number', 'whatsapp', 'cell', 'tel']
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value
        if header and any(kw in str(header).lower() for kw in keywords):
            return col
    return None


def check_whatsapp_number(driver, number):
    """
    Check if a number is on WhatsApp using WhatsApp Web.
    Navigate to the send URL and check if an error popup appears.
    Returns True if on WhatsApp, False if not.
    """
    clean = number.lstrip('+')
    url = f"https://web.whatsapp.com/send?phone={clean}"
    driver.get(url)
    time.sleep(5)  # Wait for page to load and check

    # Check for error popup: "Phone number shared via url is invalid"
    try:
        # Look for the error dialog that WhatsApp Web shows for invalid numbers
        error_selectors = [
            "//*[contains(text(), 'Phone number shared via url is invalid')]",
            "//*[contains(text(), 'phone number shared via url is invalid')]",
            "//*[contains(text(), \"isn't on WhatsApp\")]",
            "//*[contains(text(), 'not on WhatsApp')]",
        ]
        for selector in error_selectors:
            elements = driver.find_elements(By.XPATH, selector)
            if elements:
                # Click OK to dismiss the popup if there's a button
                try:
                    ok_btn = driver.find_element(By.XPATH, "//div[@role='button' and contains(text(), 'OK')]")
                    ok_btn.click()
                    time.sleep(1)
                except:
                    pass
                return False

        # Also check if the chat loaded successfully by looking for message input container
        try:
            WebDriverWait(driver, 3).until(
                EC.presence_of_element_located((By.ID, "main"))
            )
            return True  # Chat opened = number is on WhatsApp
        except:
            # No chat input found, but also no error - could be loading
            # Wait a bit more and re-check for errors
            time.sleep(3)
            for selector in error_selectors:
                elements = driver.find_elements(By.XPATH, selector)
                if elements:
                    try:
                        ok_btn = driver.find_element(By.XPATH, "//div[@role='button' and contains(text(), 'OK')]")
                        ok_btn.click()
                        time.sleep(1)
                    except:
                        pass
                    return False
            return True  # Assume present if no error found

    except Exception as e:
        print(f"    Error checking: {e}")
        return False


def main():

    #  CONFIGURE THESE VALUES BEFORE RUNNING
    filepath = "Keval's Data (LinkedIn).xlsx"          # <-- Your Excel file name here
    column_name = "P.Contact"             # <-- Column header name, or None to auto-detect
    delay = 2.0                      # <-- Seconds between checking each number

    print("\n" + "=" * 60)
    print("  WhatsApp Contact Checker (via WhatsApp Web)")
    print("=" * 60)

    # --- Load workbook ---
    if not os.path.exists(filepath):
        print(f"  Error: File not found -> {filepath}")
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active
    print(f"\n  Loading: {filepath}")
    print(f"  Sheet: '{sheet.title}' | Rows: {sheet.max_row} | Columns: {sheet.max_column}")

    # --- Find contact column ---
    col_idx = None
    if column_name:
        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=col).value
            if header and str(header).strip().lower() == column_name.strip().lower():
                col_idx = col
                break
        if col_idx is None:
            print(f"  Error: Column '{column_name}' not found.")
            sys.exit(1)
    else:
        col_idx = find_contact_column(sheet)
        if col_idx is None:
            print("  Could not auto-detect contact column.")
            for c in range(1, sheet.max_column + 1):
                print(f"    Column {c}: {sheet.cell(row=1, column=c).value}")
            sys.exit(1)

    col_header = sheet.cell(row=1, column=col_idx).value
    print(f"  Using column {col_idx}: '{col_header}'")

    # --- Find or create "Whatsapp Status" column ---
    # First, check if the column already exists anywhere in the header row
    status_col = None
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value
        if header and str(header).strip().lower() == "whatsapp status":
            status_col = col
            break
    if status_col is None:
        # Find the first truly empty column by scanning the header row
        # (avoids landing on AA/AB when max_column is inflated)
        first_empty = 1
        for col in range(1, sheet.max_column + 2):
            val = sheet.cell(row=1, column=col).value
            if val is None or str(val).strip() == "":
                first_empty = col
                break
        status_col = first_empty
        sheet.cell(row=1, column=status_col).value = "Whatsapp Status"
        print(f"  Created new column {status_col}: 'Whatsapp Status'")
    else:
        print(f"  Using existing column {status_col}: 'Whatsapp Status'")

    # --- Collect contacts ---
    contacts = []
    for row in range(2, sheet.max_row + 1):
        raw = sheet.cell(row=row, column=col_idx).value
        number = clean_number(raw)
        if number:
            contacts.append((row, number))

    print(f"  Found {len(contacts)} contact number(s).\n")

    if not contacts:
        print("  No valid contact numbers found. Exiting.")
        sys.exit(0)

    # --- Launch WhatsApp Web via Chrome browser ---
    print("  Launching WhatsApp Web in Chrome browser...")
    print("  >> If not logged in, scan the QR code with your phone.")
    print("  >> The script will wait for you to log in.\n")

    chrome_options = ChromeOptions()
    # Keep browser open after script finishes so user can review
    chrome_options.add_experimental_option("detach", True)
    
    # Use a local persistent profile so WhatsApp Web stays logged in between runs
    profile_dir = os.path.join(os.getcwd(), "chrome_profile")
    chrome_options.add_argument(f"user-data-dir={profile_dir}")
    
    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=chrome_options)
    driver.maximize_window()
    driver.get("https://web.whatsapp.com")

    # Wait until WhatsApp Web is fully loaded (look for side panel)
    print("  Waiting for WhatsApp Web to load (scan QR if needed)...")
    try:
        WebDriverWait(driver, 120).until(
            EC.presence_of_element_located((By.ID, "side"))
        )
    except:
        print("  Timeout waiting for WhatsApp Web login. Exiting.")
        driver.quit()
        sys.exit(1)

    print("  WhatsApp Web loaded!\n")
    time.sleep(2)

    # --- Check each number ---
    print("-" * 60)
    print(f"  {'Row':<6} {'Number':<18} {'Status'}")
    print("-" * 60)

    present_count = 0
    not_found_count = 0

    for i, (row, number) in enumerate(contacts, 1):
        clean = number.lstrip('+')
        is_on_whatsapp = check_whatsapp_number(driver, number)

        if is_on_whatsapp:
            status = "Present"
            present_count += 1
        else:
            status = "Not Found"
            not_found_count += 1

        sheet.cell(row=row, column=status_col).value = status
        print(f"  {row:<6} {clean:<18} {status}")

        # ===== LIVE SAVE: flush to Excel after every contact =====
        try:
            wb.save(filepath)
        except PermissionError:
            print(f"         (!) Could not save — file is open in another app. Will retry next row.")
        except Exception as e:
            print(f"         (!) Save warning: {e}")

        if i < len(contacts):
            time.sleep(delay)

    # --- Final save (in case the last live-save was skipped) ---
    try:
        wb.save(filepath)
    except Exception:
        pass
    wb.close()
    driver.quit()

    print("-" * 60)
    print(f"\n  Results: {present_count} Present, {not_found_count} Not Found")
    print(f"  Excel updated! '{filepath}' now has 'Whatsapp Status' column.")
    print(f"  Done!\n")


if __name__ == "__main__":
    main()
